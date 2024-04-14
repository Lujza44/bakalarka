'''
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

data = pd.read_csv('data/raw_vis.csv')

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(data, index=False, header=True):
    ws.append(r)

red_underlined_font = Font(color='FF0000', underline='single')


for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
    snps_5 = row[7].value  
    snps_3 = row[9].value  

    if snps_5 is not None and not pd.isna(snps_5) and str(snps_5).strip():
        snps_5_list = snps_5.split(",")
        snps_5_list = [int(num.strip()) for num in snps_5_list]
        
        #row[6].font = red_underlined_font  

    if snps_3 is not None and not pd.isna(snps_3) and str(snps_3).strip():
        snps_5_list = snps_5.split(",")
        snps_5_list = [int(num.strip()) for num in snps_5_list]
        
        #row[8].font = red_underlined_font  


wb.save('data/formatted_output.xlsx')
'''

import pandas as pd
import xlsxwriter

# Load the CSV file
data = pd.read_csv('data/raw_vis.csv')

# Create a new Excel file and add a worksheet
workbook = xlsxwriter.Workbook('data/formatted_data.xlsx')
worksheet = workbook.add_worksheet()

# Define the red font format for sequences
red_format = workbook.add_format({'font_color': 'red'})

# Write the headers
for col_num, value in enumerate(data.columns):
    worksheet.write(0, col_num, value)

# Iterate over the DataFrame rows
for row_index, row in data.iterrows():
    for col_index, value in enumerate(row):
        # Apply red format if SNP indexes columns are not empty and correspond to sequence columns
        if col_index == 6 and pd.notna(row['5\' SNP indexes']):  # 7th column (0-indexed), SNP indexes at 8th column
            worksheet.write(row_index + 1, col_index, value, red_format)
        elif col_index == 8 and pd.notna(row['3\' SNP indexes']):  # 9th column, SNP indexes at 10th column
            worksheet.write(row_index + 1, col_index, value, red_format)
        else:
            worksheet.write(row_index + 1, col_index, value)

# Close the workbook to save it
workbook.close()